import re
import json
import numpy as np
import pandas as pd
import xlrd
import openpyxl
import xlsxwriter
import schedule
import time
from matplotlib import pyplot as plt
from datetime import datetime
from openpyxl.utils.cell import get_column_letter
from tweepy.streaming import StreamListener
from tweepy import OAuthHandler
from tweepy import Stream
from tweepy import API
from tweepy import Cursor
import os



CONSUMER_KEY = "5ydgImRijCybz4oRJT9UfV4qS"
CONSUMER_SECRET = "0Nxghbj7NEXpf2B18XL6Inbhc6H4G6vGNdg40uLfmS4nmCSFu9"
ACCESS_TOKEN = "1247666477043499008-TVBjTS8nlbIPVaxUrbnI2hIEy3o5Mj"
ACCESS_TOKEN_SECRET = "rKBM9bJdDALtNjMaNjQJh7y8HXAJ4ZccatjtID3Yq9Tyi"

# # # # TWITTER CLIENT # # # #
class TwitterClient():
    def __init__(self, twitter_user=None):
        self.auth = TwitterAuthenticator().authenticate_twitter_app()
        self.twitter_client = API(self.auth)
        self.twitter_user = twitter_user
        #need to specify user, if not it will default to me
        #the parameter is none it will default to me
    def get_user_timeline_tweets(self, num_tweets):
        tweets = [] #empty list
        #id=self.twitter_user gets you user tweets of the person you sepcify
        for tweet in Cursor(self.twitter_client.user_timeline, id=self.twitter_user).items(num_tweets):
            tweets.append(tweet)
        return tweets
    def get_friend_list(self, num_friends):
        friend_list = []
        for friend in Cursor(self.twitter_client.friends, id=self.twitter_user).items(num_friends):
            friend_list.append(friend)
        return friend_list
    def get_home_timeline_tweets(self, num_tweets):
        home_timeline_tweets = []
        for tweet in Cursor(self.twitter_client.home_timeline, id=self.twitter_user).items (num_tweets):
            home_timeline_tweets.append(tweet)
        return home_timeline_tweets
    def get_twitter_client_api(self):
        return self.twitter_client

# # # # TWITTER AUTHENTICATER # # # #
class TwitterAuthenticator():
    def authenticate_twitter_app(self):
        auth = OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET)
        auth.set_access_token(ACCESS_TOKEN, ACCESS_TOKEN_SECRET)
        return auth

# # # # TWITTER STREAMER # # # #
class TwitterStreamer():
    """
    Class for streaming and processing live tweets.
    """
    def __init__(self):
        self.twitter_authenticator = TwitterAuthenticator()

    def stream_tweets(self, fetched_tweets_filename, hash_tag_list):
        # This handles Twitter authetification and the connection to Twitter Streaming API
        listener = TwitterListener(fetched_tweets_filename)
        auth = self.twitter_authenticator.authenticate_twitter_app()
        stream = Stream(auth, listener)

        # This line filter Twitter Streams to capture da ta by the keywords:
        stream.filter(track=hash_tag_list)


# # # # TWITTER STREAM LISTENER # # # #
class TwitterListener(StreamListener):
    """
    This is a basic listener that just prints received tweets to stdout.
    """
    def __init__(self, fetched_tweets_filename):
        self.fetched_tweets_filename = fetched_tweets_filename

    def on_data(self, data):
        try:
            print(data)
            with open(self.fetched_tweets_filename, 'a') as tf:
                tf.write(data)
            return True
        except BaseException as e:
            print("Error on_data %s" % str(e))
        return True


    def on_error(self, status):
        if status == 420:
            # Returning False on data method i ncase rate limit is occurs.
            return False
        print(status)



class DateEditing():
    def dateToString(self):
        source = '/Users/calebdimenstein/Desktop/Twitter/Sources/Sources_'
        date = excel_formatter.excel_handle_generator()
        xlsx = '.xlsx'
        source_day = source+date+xlsx
        file = (source_day)
        wb = xlrd.open_workbook(file)
        sheet1 = wb.sheet_by_index(0)
        testDate = sheet1.row_values(2)
        #print(testDate[4])
        #print(testDate[1])
        def floatHourToTime(fh):
            h, r = divmod(fh, 1)
            m, r = divmod(r*60, 1)
            return (
                int(h),
                int(m),
                int(r*60),
            )



        dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(testDate[4]) - 2)
        hour, minute, second = floatHourToTime(testDate[4] % 1)

        updatedDate = dt.strftime("%Y")+"-"+dt.strftime("%m")+'-'+dt.strftime("%d")
        #print(updatedDate)


        wb = openpyxl.load_workbook(source_day)
        wb.sheetnames
        sheet = wb["Sheet1"]
        amountOfRows = sheet.max_row
        amountOfColumns = sheet.max_column
        count=0;

        for k in range(1,amountOfRows):
            count=count+1
            testDate = sheet1.row_values(k)
            #print("Test Date: ",type(testDate[4]))
            #print("Test Date Value: ",testDate[4])
            if sheet1.row_values(k) ==  'date': continue
            if sheet1.row_values(k)!= 'date':
                #print("Test Date: ",type(testDate[4]))
                #print("Test Date Value: ",testDate[4])
                dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(testDate[4]) - 2)
                updatedDate = dt.strftime("%Y")+"-"+dt.strftime("%m")+'-'+dt.strftime("%d")
                cell = str(sheet[get_column_letter(5)+str(k)].value)
                #print("TESTDATE:  ",dt)
                #print("NEW DATE SHOULD MATCH  "+cell)
                #print("COUNT",count)
                sheet[get_column_letter(5)+str(k)]=updatedDate

        for k in range(1,amountOfRows):
            sheet.cell(row = k, column = 12).value = sheet.cell(row = k, column = 5).value
        for m in range (1,amountOfRows):
            sheet.cell(row = m, column = 5).value = sheet.cell(row = m, column = 13).value
        for l in range (1,amountOfRows):
            sheet.cell(row = l+1, column = 5).value = sheet.cell(row = l, column = 12).value
        for n in range (1,amountOfRows):
            sheet.cell(row = n, column = 12).value = sheet.cell(row = n, column = 14).value
        sheet.cell(row = 1, column = 5).value = "Date"
        wb.save(source_day)
        return wb
    def excel_handle_generator(self):
        dt = datetime.today()
        day = dt.day
        month = dt.month
        excel_handle = str(month)+"_"+str(day)
        return excel_handle

class TweetAnalyzer():
    """
    Functionality for analyzing and categorizing content of num_tweets
    """

    def tweets_to_data_frame(self, tweets):

        count =0
        df = pd.DataFrame(columns=['tweets','id','len','date','source','likes','retweets', 'screenName'])
        #print(df.head())

        #creating dataframe where the tweet text has to contain the mention of corona
        for tweet in tweets:
            count=count+1
            #print("Date Type: ",type(tweet.created_at))
            json_str = json.dumps(tweet._json)
            parsed = json.loads(json_str)
            #print(json.dumps(parsed, indent=4, sort_keys=True))
            #print(parsed['text'])
            #print('COUNT: ',count)
            u=re.findall('.+Corona.+',parsed['text'])
            v=re.findall('.+Covid.+',parsed['text'])
            w=re.findall('.+corona.+',parsed['text'])
            x=re.findall('.+virus.+',parsed['text'])
            y=re.findall('.+pandemic.+',parsed['text'])
            z=re.findall('.+COVID.+',parsed['text'])
            if ((len(u) > 0 or len(v) > 0 or len(w) > 0 or len(x) > 0 or len(y) > 0 or len(z) > 0)):
                df = df.append({'tweets':parsed['text'], 'id':tweet.id, 'len':len(tweet.text), 'date':tweet.created_at, 'source':tweet.source,'likes':tweet.favorite_count, 'retweets':tweet.retweet_count, 'screenName':parsed['user']['screen_name']},ignore_index=True)


        return df

if __name__ == '__main__':
    excel_formatter = DateEditing()
    source = '/Users/calebdimenstein/Desktop/Twitter/Sources/Sources_'
    date = excel_formatter.excel_handle_generator()
    xlsx = '.xlsx'
    source_day = source+date+xlsx


    workbook_filePath = source_day
    wb = openpyxl.Workbook()
    wb.save(workbook_filePath)

    twitter_client = TwitterClient()
    api = twitter_client.get_twitter_client_api()
    tweet_analyzer = TweetAnalyzer()

    twitter_client1 = TwitterClient()
    api1 = twitter_client1.get_twitter_client_api()
    tweet_analyzer1 = TweetAnalyzer()

    #string array of Twitter handles that I run through, put each mention of COVID into a DB, merge them one at a time into a larger DB, then export on excel
    twitter_names = ["nytimes", "latimes", "WSJ", "CNN", "NyGovCuomo", "SenSanders", "BarackObama", "MSNBC", "BostonGlobe", "JoeBiden", "foxnewsalert", "KylieJenner", "Forbes", "TheEconomist", "BBCWorld", "NewYorker", "TheAtlantic", "politico", "cnnbrk", "jimmyfallon", "StephenAtHome", "TheDailyShow", "Trevornoah", "LateNightSeth"]
    dataframes = []
    #Loop that goes through each Twitter user
    for name in twitter_names:
        tweets = api.user_timeline(screen_name =name, count=200)
        df = tweet_analyzer.tweets_to_data_frame(tweets)
        dataframes.append(df)
    results=pd.concat(dataframes, ignore_index=True)
    results.to_excel(source_day)
    print (results)
    print(results.size)

    #fixes the dates
    print("DATE EDITOR RUN")
    date_editor = DateEditing()
    newDates = date_editor.dateToString()




    # average length of the tweets
    #print(np.mean(df['len']))
    # likes that the tweet that recieved most likes
    #print (np.max(df['likes']))
    # Number of retweets for the most retweeted
    #print (np.max(df['retweet']))

    #Time Series (index is x axis and likes is y axis)
    #time_likes = pd.Series(data=df['likes'].values, index=df['date'])
    #time_likes.plot(figsize=(16,4), color='r')
    #plt.show()

    #Time series for retweets
    #time_retweets = pd.Series(data=df['retweets'].values, index=df['date'])
    #time_retweets.plot(figsize=(16,4), color='r')
    #plt.show()


    #
