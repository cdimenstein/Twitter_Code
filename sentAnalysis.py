import xlrd
import openpyxl
import nltk
from openpyxl.utils.cell import get_column_letter
from nltk.sentiment.vader import SentimentIntensityAnalyzer
import numpy as np
import pandas as pd
nltk.download('vader_lexicon')

def nltk_sentiment(sentence):
    nltk_sentiment = SentimentIntensityAnalyzer()
    score = nltk_sentiment.polarity_scores(sentence)
    return score

wb = openpyxl.load_workbook('/Users/calebdimenstein/Desktop/Twitter/Updated_Sources.xlsx')

sheet1 = wb["Sheet1"]

amountOfRows1 = 4974
amountOfColumns1 = sheet1.max_column

tweet_list = []

for r in range(1,amountOfRows1):
    tweet_list.append(sheet1.cell(row = r, column = 2).value)

nltk_results = [nltk_sentiment(row) for row in tweet_list]
results_df = pd.DataFrame(nltk_results)
text_df = pd.DataFrame(tweet_list, columns = ['text'])
nltk_df = text_df.join(results_df)
print(nltk_df)
nltk_df.to_excel('/Users/calebdimenstein/Desktop/test_run.xlsx')
